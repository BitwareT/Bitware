# -*- coding: utf-8 -*-
from flask import Flask, request, jsonify
from flask_cors import CORS
import mysql.connector
import random
import re
from datetime import datetime
import os
import uuid
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.styles import Font, PatternFill

# --- NUEVAS IMPORTACIONES PARA ARREGLAR ERRORES Y WARNINGS ---
from sqlalchemy import create_engine
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.naive_bayes import MultinomialNB
from statsmodels.tsa.statespace.sarimax import SARIMAX
import warnings

app = Flask(__name__, static_folder='static')
CORS(app)
app.json.ensure_ascii = False
app.config['JSON_AS_ASCII'] = False

# --- CONFIGURACIíN DE DIRECTORIOS ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXPORT_DIR = os.path.join(BASE_DIR, 'static', 'exports')
os.makedirs(EXPORT_DIR, exist_ok=True)

# --- CONEXIíN SQLALCHEMY (CORRIGE ERROR DE PANDAS) ---
# Usaremos esto solo cuando usemos pandas (Predicciones y Excel)
DB_CONNECTION_STR = 'mysql+pymysql://bitware_user:Rocky25..@localhost/bitware'
db_engine = create_engine(DB_CONNECTION_STR)

# --- CONEXIíN CLíSICA (Para el resto del sistema) ---
def conectar_db():
    try:
        return mysql.connector.connect(host="localhost", user="bitware_user", password="Rocky25..", database="bitware")
    except mysql.connector.Error as err:
        print(f"Error de Base de Datos: {err}")
        return None

# ======================================================================
# ENTRENAMIENTO DEL MODELO DE INTELIGENCIA ARTIFICIAL (CHATBOT)
# ======================================================================

# --- Intenciones de Usuario ---
saludos = ["hola", "buenas", "qué tal", "hey", "saludos"]
productos_frases = [
    "componentes", "productos", "recomiéndame algo", "hardware", "accesorios", "recomendaciones",
    "recomendar producto", "recomiéndame un producto", "dame una recomendación", "muéstrame productos"
]
pedidos_frases = ["mi pedido", "mi orden", "mi compra", "dónde está mi pedido", "estado de mi orden"]
soportes_frases = ["soporte", "ayuda", "problema técnico", "mi pc no prende", "pantalla azul"]
funciones_frases = ["funciones", "qué puedes hacer", "qué haces", "cuáles son tus funciones", "menú de opciones"]
busqueda_frases = ["busca", "búscame", "encuentra", "tienes", "precio de", "cotiza", "buscar producto", "quiero buscar"]
devolucion_frases = ["devolucion", "quiero devolver", "iniciar una devolución", "mi producto llegó dañado", "devolver pedido"]
notificacion_stock_frases = ["aví­same cuando llegue", "notifí­came del stock", "cuando vuelve a estar disponible"]
comparar_frases = [
    "compara", "cuál es mejor", "diferencia entre", 
    "compara a con b", "comparar", "comparativa", "vs",
    "compara este producto con este otro"
]
actualizar_direccion_frases = ["actualizar mi dirección", "cambiar dirección de enví­o", "modificar dirección"]

# --- Intenciones de Administrador ---
stats_admin_frases = ["estadí­sticas", "stats", "resumen", "reporte", "cómo van las ventas", "total usuarios", "reporte de ventas hoy"]
stock_admin_frases = ["stock de", "cuánto stock queda de", "inventario de", "revisar stock", "consultar stock"]
buscar_cliente_frases = ["busca al cliente", "datos de cliente", "info de", "quién es el cliente", "encuentra a", "cliente"]
cambiar_estado_pedido_frases = ["cambia el estado del pedido", "actualiza el pedido", "marcar como enviado"]
analisis_frases = ["análisis", "analisis", "análisis de crecimiento", "qué categorí­a vende más", "clientes inactivos"]
prediccion_frases = ["predice el stock", "predicción de", "pronostica", "cuánto se venderá de", "demanda de", "predecir stock"]

# ---- Intenciones Invitados ----
horarios_frases = ["horarios de atención", "cuál es su horario", "a qué hora abren", "atienden los sábados", "horario"]
pagos_frases = ["métodos de pago", "cómo puedo pagar", "aceptan tarjeta", "se puede pagar con transferencia", "formas de pago"]

# ---- Intenciones Vendedor (NUEVO) ----
exportar_frases = ["exportar ventas", "descargar ventas", "dame un excel con mis ventas", "generar excel", "ventas en excel", "reporte excel"]

# Combinamos todas las frases e intenciones
frases = (saludos + productos_frases + pedidos_frases + soportes_frases + funciones_frases + busqueda_frases +
          devolucion_frases + notificacion_stock_frases + comparar_frases + actualizar_direccion_frases +
          horarios_frases + pagos_frases +
          stats_admin_frases + stock_admin_frases + buscar_cliente_frases + cambiar_estado_pedido_frases + analisis_frases +
          prediccion_frases + exportar_frases)

intenciones = (
    ["saludo"] * len(saludos) + ["producto"] * len(productos_frases) + ["pedido"] * len(pedidos_frases) +
    ["soporte"] * len(soportes_frases) + ["funciones"] * len(funciones_frases) + ["busqueda_producto"] * len(busqueda_frases) +
    ["solicitar_devolucion"] * len(devolucion_frases) + ["solicitar_notificacion"] * len(notificacion_stock_frases) +
    ["comparar_productos"] * len(comparar_frases) + ["actualizar_direccion"] * len(actualizar_direccion_frases) +
    ["horarios"] * len(horarios_frases) + ["pagos"] * len(pagos_frases) +
    ["stats_admin"] * len(stats_admin_frases) + ["stock_admin"] * len(stock_admin_frases) +
    ["buscar_cliente_admin"] * len(buscar_cliente_frases) + ["cambiar_estado_pedido"] * len(cambiar_estado_pedido_frases) +
    ["analisis_admin"] * len(analisis_frases) +
    ["prediccion_stock"] * len(prediccion_frases) +
    ["exportar_ventas"] * len(exportar_frases)
)

vectorizer = CountVectorizer()
X = vectorizer.fit_transform(frases)
modelo = MultinomialNB()
modelo.fit(X, intenciones)

def clasificar_intencion(mensaje):
    X_new = vectorizer.transform([mensaje])
    return modelo.predict(X_new)[0]

# ======================================================================
# FUNCIONES DE BASE DE DATOS (Chatbot y Predicción)
# ======================================================================

def recomendar_productos():
    conn = conectar_db()
    if not conn: return []
    try:
        cursor = conn.cursor(dictionary=True)
        query = "SELECT id_producto AS id, nombre, precio, imagen_principal FROM producto WHERE stock > 0 AND imagen_principal IS NOT NULL AND activo = 1 ORDER BY RAND() LIMIT 3"
        cursor.execute(query)
        productos = cursor.fetchall()
        for p in productos:
            if p.get('precio'): p['precio'] = float(p['precio'])
        return productos
    finally:
        if conn and conn.is_connected(): conn.close()

def buscar_productos_por_nombre(termino_busqueda):
    conn = conectar_db()
    if not conn: return []
    try:
        cursor = conn.cursor(dictionary=True)
        query = """
            SELECT id_producto AS id, nombre, precio, imagen_principal
            FROM producto
            WHERE (nombre LIKE %s OR categoria LIKE %s)
              AND activo = 1
            LIMIT 3
        """
        like_term = f"%{termino_busqueda}%"
        cursor.execute(query, (like_term, like_term))
        productos = cursor.fetchall()
        processed_productos = []
        for p in productos:
            try:
                if p.get('precio') is not None:
                    p['precio'] = float(p['precio'])
                processed_productos.append(p)
            except Exception as e:
                print(f"!!! ERROR: Procesando producto {p.get('id')}: {e}")
        return processed_productos
    except Exception as e:
        print(f"!!! ERROR inesperado en buscar_productos_por_nombre: {e}")
        return []
    finally:
        if conn and conn.is_connected():
            conn.close()

# --- NUEVA FUNCIíN: EXCEL (USANDO SQLALCHEMY) ---
def generar_excel_ventas(id_vendedor, base_url, es_admin=False):
    try:
        # --- 1. CONSULTA SQL ---
        if es_admin:
            query = """
                SELECT 
                    DATE(p.fecha_pedido) AS 'Fecha',
                    p.id_pedido AS 'ID Pedido',
                    pr.nombre AS 'Producto',
                    pr.categoria AS 'Categoria', 
                    pp.cantidad AS 'Cantidad',
                    pp.precio_unitario AS 'Precio Unitario',
                    (pp.cantidad * pp.precio_unitario) AS 'Total Venta',
                    p.estado AS 'Estado Pedido',
                    u.nombre AS 'Cliente',
                    u.email AS 'Email Cliente',
                    u.region AS 'Region',
                    pr.id_vendedor AS 'ID Vendedor'
                FROM pedidos_productos pp
                JOIN pedidos p ON pp.id_pedido = p.id_pedido
                JOIN producto pr ON pp.id_producto = pr.id_producto
                JOIN usuario u ON p.id_usuario = u.id_usuario
                WHERE p.estado IN ('Pagado', 'Enviado', 'Entregado')
                ORDER BY p.fecha_pedido DESC
            """
            params = None
        else:
            query = """
                SELECT 
                    DATE(p.fecha_pedido) AS 'Fecha',
                    p.id_pedido AS 'ID Pedido',
                    pr.nombre AS 'Producto',
                    pr.categoria AS 'Categoria',
                    pp.cantidad AS 'Cantidad',
                    pp.precio_unitario AS 'Precio Unitario',
                    (pp.cantidad * pp.precio_unitario) AS 'Total Venta',
                    p.estado AS 'Estado Pedido',
                    u.nombre AS 'Cliente',
                    u.email AS 'Email Cliente',
                    u.region AS 'Region'
                FROM pedidos_productos pp
                JOIN pedidos p ON pp.id_pedido = p.id_pedido
                JOIN producto pr ON pp.id_producto = pr.id_producto
                JOIN usuario u ON p.id_usuario = u.id_usuario
                WHERE pr.id_vendedor = %s
                  AND p.estado IN ('Pagado', 'Enviado', 'Entregado')
                ORDER BY p.fecha_pedido DESC
            """
            params = (id_vendedor,)

        # --- 2. PROCESAR DATOS ---
        df = pd.read_sql(query, db_engine, params=params)
        if df.empty: return "empty"

        fecha_hora = datetime.now().strftime("%Y-%m-%d_%H-%M")
        prefix = "Reporte_Global" if es_admin else f"Ventas_Vendedor_{id_vendedor}"
        filename = f"{prefix}_{fecha_hora}.xlsx"
        filepath = os.path.join(EXPORT_DIR, filename)

        # --- 3. CREAR TABLAS RESUMEN ---
        # Tabla A: Productos (Columna A=1, B=2)
        res_prod = df.groupby('Producto')[['Total Venta']].sum().sort_values('Total Venta', ascending=False).reset_index()
        # Tabla B: Categorías (Columna D=4, E=5)
        res_cat = df.groupby('Categoria')[['Total Venta']].sum().reset_index()
        # Tabla C: Regiones (Columna G=7, H=8)
        res_reg = df.groupby('Region')[['Total Venta']].sum().sort_values('Total Venta', ascending=False).head(10).reset_index()
        # Tabla D: Fechas (Columna J=10, K=11)
        res_date = df.groupby('Fecha')[['Total Venta']].sum().reset_index()

        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Detalle', index=False)
            
            # Ubicamos las tablas separadas por una columna vacía
            res_prod.to_excel(writer, sheet_name='Dashboard', startcol=0, index=False)  # Cols A, B
            res_cat.to_excel(writer, sheet_name='Dashboard', startcol=3, index=False)   # Cols D, E
            res_reg.to_excel(writer, sheet_name='Dashboard', startcol=6, index=False)   # Cols G, H
            res_date.to_excel(writer, sheet_name='Dashboard', startcol=9, index=False)  # Cols J, K

        # --- 4. GENERAR GRÁFICOS (CORREGIDO) ---
        wb = load_workbook(filepath)
        ws = wb['Dashboard']
        
        # --- GRÁFICO 1: BARRAS (Productos) ---
        # Datos en Columna 2 (B)
        chart1 = BarChart()
        chart1.type = "col"
        chart1.style = 10
        chart1.title = "Top Productos (Ingresos)"
        chart1.y_axis.title = "$"
        chart1.x_axis.title = "Producto"
        
        filas_prod = len(res_prod) + 1
        data1 = Reference(ws, min_col=2, min_row=1, max_row=filas_prod, max_col=2)
        cats1 = Reference(ws, min_col=1, min_row=2, max_row=filas_prod)
        chart1.add_data(data1, titles_from_data=True)
        chart1.set_categories(cats1)
        chart1.width = 18
        chart1.height = 10
        ws.add_chart(chart1, "A20") 

        # --- GRÁFICO 2: TORTA (Categorías) ---
        # CORRECCIÓN: Datos en Columna 5 (E), Categorías en 4 (D)
        chart2 = PieChart()
        chart2.title = "Ventas por Categoría"
        
        filas_cat = len(res_cat) + 1
        data2 = Reference(ws, min_col=5, min_row=1, max_row=filas_cat, max_col=5) # <--- CORREGIDO (Era 4)
        cats2 = Reference(ws, min_col=4, min_row=2, max_row=filas_cat) # <--- CORREGIDO (Era 3)
        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(cats2)
        ws.add_chart(chart2, "F20") 

        # --- GRÁFICO 3: TORTA (Regiones) ---
        # CORRECCIÓN: Datos en Columna 8 (H), Categorías en 7 (G)
        chart3 = PieChart()
        chart3.title = "Ventas por Región"
        
        filas_reg = len(res_reg) + 1
        data3 = Reference(ws, min_col=8, min_row=1, max_row=filas_reg, max_col=8) # <--- CORREGIDO (Era 7)
        cats3 = Reference(ws, min_col=7, min_row=2, max_row=filas_reg) # <--- CORREGIDO (Era 6)
        chart3.add_data(data3, titles_from_data=True)
        chart3.set_categories(cats3)
        ws.add_chart(chart3, "K20")

        # --- GRÁFICO 4: LÍNEA (Tendencia) ---
        # CORRECCIÓN: Datos en Columna 11 (K), Categorías en 10 (J)
        chart4 = LineChart()
        chart4.title = "Tendencia de Ventas (Diaria)"
        chart4.style = 12
        chart4.y_axis.title = "$"
        chart4.x_axis.title = "Fecha"
        
        filas_date = len(res_date) + 1
        data4 = Reference(ws, min_col=11, min_row=1, max_row=filas_date, max_col=11) # <--- CORREGIDO (Era 10)
        cats4 = Reference(ws, min_col=10, min_row=2, max_row=filas_date) # <--- CORREGIDO (Era 9)
        chart4.add_data(data4, titles_from_data=True)
        chart4.set_categories(cats4)
        chart4.width = 30
        chart4.height = 10
        ws.add_chart(chart4, "A40")

        wb.save(filepath)
        return f"{base_url}static/exports/{filename}"

    except Exception as e:
        print(f"Error generando Excel Dashboard: {e}")
        return None
    try:
        # --- 1. CONSULTA SQL (Agregamos 'pr.categoria' para el nuevo gráfico) ---
        if es_admin:
            query = """
                SELECT 
                    DATE(p.fecha_pedido) AS 'Fecha',
                    p.id_pedido AS 'ID Pedido',
                    pr.nombre AS 'Producto',
                    pr.categoria AS 'Categoria', 
                    pp.cantidad AS 'Cantidad',
                    pp.precio_unitario AS 'Precio Unitario',
                    (pp.cantidad * pp.precio_unitario) AS 'Total Venta',
                    p.estado AS 'Estado Pedido',
                    u.nombre AS 'Cliente',
                    u.email AS 'Email Cliente',
                    u.region AS 'Region',
                    pr.id_vendedor AS 'ID Vendedor'
                FROM pedidos_productos pp
                JOIN pedidos p ON pp.id_pedido = p.id_pedido
                JOIN producto pr ON pp.id_producto = pr.id_producto
                JOIN usuario u ON p.id_usuario = u.id_usuario
                WHERE p.estado IN ('Pagado', 'Enviado', 'Entregado')
                ORDER BY p.fecha_pedido DESC
            """
            params = None
        else:
            query = """
                SELECT 
                    DATE(p.fecha_pedido) AS 'Fecha',
                    p.id_pedido AS 'ID Pedido',
                    pr.nombre AS 'Producto',
                    pr.categoria AS 'Categoria',
                    pp.cantidad AS 'Cantidad',
                    pp.precio_unitario AS 'Precio Unitario',
                    (pp.cantidad * pp.precio_unitario) AS 'Total Venta',
                    p.estado AS 'Estado Pedido',
                    u.nombre AS 'Cliente',
                    u.email AS 'Email Cliente',
                    u.region AS 'Region'
                FROM pedidos_productos pp
                JOIN pedidos p ON pp.id_pedido = p.id_pedido
                JOIN producto pr ON pp.id_producto = pr.id_producto
                JOIN usuario u ON p.id_usuario = u.id_usuario
                WHERE pr.id_vendedor = %s
                  AND p.estado IN ('Pagado', 'Enviado', 'Entregado')
                ORDER BY p.fecha_pedido DESC
            """
            params = (id_vendedor,)

        # --- 2. PROCESAR DATOS ---
        df = pd.read_sql(query, db_engine, params=params)
        if df.empty: return "empty"

        fecha_hora = datetime.now().strftime("%Y-%m-%d_%H-%M")
        prefix = "Reporte_Global" if es_admin else f"Ventas_Vendedor_{id_vendedor}"
        filename = f"{prefix}_{fecha_hora}.xlsx"
        filepath = os.path.join(EXPORT_DIR, filename)

        # --- 3. CREAR TABLAS RESUMEN PARA LOS GRÁFICOS ---
        # A. Por Producto
        res_prod = df.groupby('Producto')[['Total Venta']].sum().sort_values('Total Venta', ascending=False).reset_index()
        # B. Por Categoría
        res_cat = df.groupby('Categoria')[['Total Venta']].sum().reset_index()
        # C. Por Región (Top 10 para no saturar)
        res_reg = df.groupby('Region')[['Total Venta']].sum().sort_values('Total Venta', ascending=False).head(10).reset_index()
        # D. Por Fecha (Tendencia)
        res_date = df.groupby('Fecha')[['Total Venta']].sum().reset_index()

        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Detalle', index=False)
            
            # Guardamos las tablas en la hoja Dashboard en diferentes columnas
            res_prod.to_excel(writer, sheet_name='Dashboard', startcol=0, index=False)  # Col A
            res_cat.to_excel(writer, sheet_name='Dashboard', startcol=3, index=False)   # Col D
            res_reg.to_excel(writer, sheet_name='Dashboard', startcol=6, index=False)   # Col G
            res_date.to_excel(writer, sheet_name='Dashboard', startcol=9, index=False)  # Col J

        # --- 4. GENERAR GRÁFICOS CON OPENPYXL ---
        wb = load_workbook(filepath)
        ws = wb['Dashboard']
        
        # --- GRÁFICO 1: BARRAS (Top Productos) ---
        chart1 = BarChart()
        chart1.type = "col"
        chart1.style = 10
        chart1.title = "Top Productos (Ingresos)"
        chart1.y_axis.title = "$"
        chart1.x_axis.title = "Producto"
        
        filas_prod = len(res_prod) + 1
        data1 = Reference(ws, min_col=2, min_row=1, max_row=filas_prod, max_col=2)
        cats1 = Reference(ws, min_col=1, min_row=2, max_row=filas_prod)
        chart1.add_data(data1, titles_from_data=True)
        chart1.set_categories(cats1)
        chart1.width = 18
        chart1.height = 10
        ws.add_chart(chart1, "A20") # Ubicación: Debajo de las tablas

        # --- GRÁFICO 2: TORTA (Categorías) ---
        chart2 = PieChart()
        chart2.title = "Ventas por Categoría"
        
        filas_cat = len(res_cat) + 1
        data2 = Reference(ws, min_col=4, min_row=1, max_row=filas_cat, max_col=4)
        cats2 = Reference(ws, min_col=3, min_row=2, max_row=filas_cat)
        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(cats2)
        ws.add_chart(chart2, "F20") 

        # --- GRÁFICO 3: TORTA (Regiones) ---
        chart3 = PieChart()
        chart3.title = "Ventas por Región"
        
        filas_reg = len(res_reg) + 1
        data3 = Reference(ws, min_col=7, min_row=1, max_row=filas_reg, max_col=7)
        cats3 = Reference(ws, min_col=6, min_row=2, max_row=filas_reg)
        chart3.add_data(data3, titles_from_data=True)
        chart3.set_categories(cats3)
        ws.add_chart(chart3, "K20")

        # --- GRÁFICO 4: LÍNEA (Tendencia Diaria) ---
        chart4 = LineChart()
        chart4.title = "Tendencia de Ventas (Diaria)"
        chart4.style = 12
        chart4.y_axis.title = "$"
        chart4.x_axis.title = "Fecha"
        
        filas_date = len(res_date) + 1
        data4 = Reference(ws, min_col=10, min_row=1, max_row=filas_date, max_col=10)
        cats4 = Reference(ws, min_col=9, min_row=2, max_row=filas_date)
        chart4.add_data(data4, titles_from_data=True)
        chart4.set_categories(cats4)
        chart4.width = 30 # Más ancho para ver bien las fechas
        chart4.height = 10
        ws.add_chart(chart4, "A40") # Ubicación: Más abajo

        wb.save(filepath)
        return f"{base_url}static/exports/{filename}"

    except Exception as e:
        print(f"Error generando Excel Dashboard: {e}")
        return None
    try:
        # --- 1. PREPARAR LA CONSULTA SQL ---
        if es_admin:
            query = """
                SELECT 
                    p.fecha_pedido AS 'Fecha',
                    p.id_pedido AS 'ID Pedido',
                    pr.nombre AS 'Producto',
                    pp.cantidad AS 'Cantidad',
                    pp.precio_unitario AS 'Precio Unitario',
                    (pp.cantidad * pp.precio_unitario) AS 'Total Venta',
                    p.estado AS 'Estado Pedido',
                    u.nombre AS 'Cliente',
                    u.email AS 'Email Cliente',
                    u.region AS 'Region',
                    pr.id_vendedor AS 'ID Vendedor'
                FROM pedidos_productos pp
                JOIN pedidos p ON pp.id_pedido = p.id_pedido
                JOIN producto pr ON pp.id_producto = pr.id_producto
                JOIN usuario u ON p.id_usuario = u.id_usuario
                WHERE p.estado IN ('Pagado', 'Enviado', 'Entregado')
                ORDER BY p.fecha_pedido DESC
            """
            params = None
        else:
            query = """
                SELECT 
                    p.fecha_pedido AS 'Fecha',
                    p.id_pedido AS 'ID Pedido',
                    pr.nombre AS 'Producto',
                    pp.cantidad AS 'Cantidad',
                    pp.precio_unitario AS 'Precio Unitario',
                    (pp.cantidad * pp.precio_unitario) AS 'Total Venta',
                    p.estado AS 'Estado Pedido',
                    u.nombre AS 'Cliente',
                    u.email AS 'Email Cliente',
                    u.region AS 'Región'
                FROM pedidos_productos pp
                JOIN pedidos p ON pp.id_pedido = p.id_pedido
                JOIN producto pr ON pp.id_producto = pr.id_producto
                JOIN usuario u ON p.id_usuario = u.id_usuario
                WHERE pr.id_vendedor = %s
                  AND p.estado IN ('Pagado', 'Enviado', 'Entregado')
                ORDER BY p.fecha_pedido DESC
            """
            params = (id_vendedor,)

        # --- 2. OBTENER DATOS CON PANDAS ---
        df = pd.read_sql(query, db_engine, params=params)
        if df.empty: return "empty"

        # --- 3. GENERAR NOMBRE DE ARCHIVO LIMPIO (FECHA Y HORA) ---
        # Formato: Año-Mes-Dia_Hora-Minuto (Ej: 2025-12-05_18-30)
        fecha_hora = datetime.now().strftime("%Y-%m-%d_%H-%M")
        
        if es_admin:
            filename = f"Reporte_Global_{fecha_hora}.xlsx"
        else:
            # Incluimos el ID del vendedor para diferenciar
            filename = f"Ventas_Vendedor_{id_vendedor}_{fecha_hora}.xlsx"
            
        filepath = os.path.join(EXPORT_DIR, filename)

        # --- 4. CREAR EXCEL CON HOJAS Y GRÁFICOS ---
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Hoja 1: Detalle completo
            df.to_excel(writer, sheet_name='Detalle', index=False)
            
            # Hoja 2: Resumen para el gráfico
            df_resumen = df.groupby('Producto')[['Total Venta', 'Cantidad']].sum().reset_index()
            df_resumen = df_resumen.sort_values(by='Total Venta', ascending=False)
            df_resumen.to_excel(writer, sheet_name='Resumen', index=False)

        # --- 5. AGREGAR EL GRÁFICO CON OPENPYXL ---
        wb = load_workbook(filepath)
        ws_resumen = wb['Resumen']
        
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Total de Ventas por Producto"
        chart.y_axis.title = "Ingresos ($)"
        chart.x_axis.title = "Producto"
        
        filas = len(df_resumen) + 1
        
        # Datos (Columna B: Total Venta)
        data = Reference(ws_resumen, min_col=2, min_row=1, max_row=filas, max_col=2)
        # Categorías (Columna A: Nombres de Productos)
        cats = Reference(ws_resumen, min_col=1, min_row=2, max_row=filas)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        chart.width = 20 # Hacer el gráfico más ancho
        chart.height = 10 # Hacer el gráfico más alto
        
        ws_resumen.add_chart(chart, "E2")
        
        wb.save(filepath)

        # Devolver URL de descarga
        download_url = f"{base_url}static/exports/{filename}"
        return download_url

    except Exception as e:
        print(f"Error generando Excel: {e}")
        return None
# //////////////////////////////////////////////////////////////////////////////
# ///////////////             DEVOLUCION                 ///////////////////////
# //////////////////////////////////////////////////////////////////////////////

def solicitar_devolucion_db(id_usuario):
    conn = conectar_db()
    if not conn: 
        return {"elegible": False, "mensaje": "No pude conectarme a la base de datos para verificar tus pedidos."}
    try:
        cursor = conn.cursor(dictionary=True)
        query_elegible = """
            SELECT id_pedido
            FROM pedidos 
            WHERE id_usuario = %s 
              AND estado IN ('Entregado', 'Completado')
            LIMIT 1
        """
        cursor.execute(query_elegible, (id_usuario,))
        pedido_elegible = cursor.fetchone()
        
        if pedido_elegible:
            return {
                "elegible": True, 
                "mensaje": (
                    "¡Claro! He verificado que tienes pedidos **entregados** que son elegibles para devolución.\n\n"
                    "Para iniciar la solicitud de forma segura (y adjuntar fotos si es necesario), "
                    "por favor ve a **Mi Cuenta > Mis Pedidos**.\n\n"
                    "Ahí­ verás el botón **'Solicitar Devolución'** junto a los pedidos que aplican."
                )
            }
        
        query_otros = "SELECT estado FROM pedidos WHERE id_usuario = %s ORDER BY fecha_pedido DESC LIMIT 1"
        cursor.execute(query_otros, (id_usuario,))
        ultimo_pedido = cursor.fetchone()
        
        if ultimo_pedido:
            return {
                "elegible": False, 
                "mensaje": f"Revisé tu cuenta y veo que tu último pedido está en estado **'{ultimo_pedido['estado']}'**. \n\nSolo puedes iniciar una devolución **después de que el pedido haya sido 'Entregado'**."
            }
        else:
            return {
                "elegible": False, 
                "mensaje": "Revisé tu cuenta, pero no encontré ningún pedido registrado."
            }
    except Exception as e:
        print(f"!!! ERROR en solicitar_devolucion_db: {e}")
        return {"elegible": False, "mensaje": "Tuve un problema al consultar tus pedidos."}
    finally:
        if conn and conn.is_connected(): conn.close()

def solicitar_notificacion_db(id_usuario, email_usuario, nombre_producto):
    conn = conectar_db()
    if not conn: return "No pude procesar tu solicitud."
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT id_producto, stock, nombre FROM producto WHERE nombre LIKE %s LIMIT 1", (f"%{nombre_producto}%",))
        producto = cursor.fetchone()
        if not producto:
            return f"No encontré el producto '{nombre_producto}'."
        if producto['stock'] > 0:
            return f"¡Buenas noticias! El producto '{producto['nombre']}' ya se encuentra en stock."
        id_producto = producto['id_producto']
        cursor.execute("INSERT INTO notificaciones_stock (id_usuario, id_producto, email_usuario) VALUES (%s, %s, %s)", (id_usuario, id_producto, email_usuario))
        conn.commit()
        return f"¡Entendido! Te enviaré un correo a {email_usuario} tan pronto como '{producto['nombre']}' vuelva a estar disponible."
    finally:
        if conn and conn.is_connected(): conn.close()

def comparar_productos_db(producto1_nombre, producto2_nombre):
    conn = conectar_db()
    if not conn: return "No pude obtener la información de los productos."
    try:
        cursor = conn.cursor(dictionary=True)
        query = "SELECT nombre, precio, descripcion FROM producto WHERE nombre LIKE %s OR nombre LIKE %s LIMIT 2"
        cursor.execute(query, (f"%{producto1_nombre}%", f"%{producto2_nombre}%"))
        productos = cursor.fetchall()
        if len(productos) < 2:
            return "No pude encontrar uno o ambos productos para comparar."
        p1, p2 = productos[0], productos[1]
        respuesta = (f"**Comparando {p1['nombre']} vs {p2['nombre']}:**\n"
                     f"- **Precio {p1['nombre']}**: ${p1['precio']:,.0f}\n"
                     f"- **Precio {p2['nombre']}**: ${p2['precio']:,.0f}\n")
        if p1['precio'] < p2['precio']:
            respuesta += f"**Conclusión:** {p1['nombre']} es más económico."
        else:
            respuesta += f"**Conclusión:** {p2['nombre']} es más económico."
        return respuesta
    finally:
        if conn and conn.is_connected(): conn.close()

def actualizar_direccion_db(id_usuario, nueva_direccion):
    conn = conectar_db()
    if not conn: return "No pude actualizar tu dirección."
    try:
        cursor = conn.cursor()
        cursor.execute("UPDATE usuario SET direccion = %s WHERE id_usuario = %s", (nueva_direccion, id_usuario))
        conn.commit()
        return "¡Listo! He actualizado tu dirección de enví­o principal."
    finally:
        if conn and conn.is_connected(): conn.close()
        
def estado_ultimo_pedido(id_usuario):
    conn = conectar_db()
    if not conn: return None
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT id_pedido, fecha_pedido, estado FROM pedidos WHERE id_usuario = %s ORDER BY fecha_pedido DESC LIMIT 1", (id_usuario,))
        return cursor.fetchone()
    finally:
        if conn and conn.is_connected(): conn.close()

# --- Funciones de Administrador ---
def get_proactive_alerts():
    conn = conectar_db()
    if not conn: return ""
    try:
        cursor = conn.cursor()
        alerts = []
        cursor.execute("SELECT COUNT(*) FROM producto WHERE stock < 10 AND activo = 1")
        low_stock = cursor.fetchone()[0]
        if low_stock > 0: alerts.append(f"Tienes **{low_stock} productos con bajo stock**.")
        cursor.execute("SELECT COUNT(*) FROM solicitudes_servicio WHERE estado = 'Pendiente'")
        pending_services = cursor.fetchone()[0]
        if pending_services > 0: alerts.append(f"Hay **{pending_services} solicitudes de servicio** pendientes.")
        return " ".join(alerts)
    finally:
        if conn and conn.is_connected(): conn.close()

def cambiar_estado_pedido_db(id_pedido, nuevo_estado):
    conn = conectar_db()
    if not conn: return "No pude actualizar el pedido."
    try:
        cursor = conn.cursor()
        cursor.execute("UPDATE pedidos SET estado = %s WHERE id_pedido = %s", (nuevo_estado, id_pedido))
        if cursor.rowcount > 0:
            conn.commit()
            return f"¡Hecho! El pedido #{id_pedido} ha sido actualizado a **{nuevo_estado}**."
        else:
            return f"No encontré el pedido #{id_pedido}."
    finally:
        if conn and conn.is_connected(): conn.close()
        
def get_category_growth_analysis():
    conn = conectar_db()
    if not conn: return "No pude realizar el análisis."
    try:
        cursor = conn.cursor(dictionary=True)
        query = """
            SELECT pr.categoria, SUM(p.total) as ventas_mes_actual
            FROM pedidos p JOIN producto pr ON p.id_producto = pr.id_producto
            WHERE p.estado = 'Pagado' AND p.fecha_pedido >= DATE_FORMAT(NOW(), '%Y-%m-01')
            GROUP BY pr.categoria ORDER BY ventas_mes_actual DESC LIMIT 1;
        """
        cursor.execute(query)
        top_category = cursor.fetchone()
        if not top_category:
            return "No hay ventas este mes para analizar."
        return f"La categorí­a con mayores ingresos este mes es **{top_category['categoria'].upper()}**."
    finally:
        if conn and conn.is_connected(): conn.close()

def buscar_cliente_por_email_o_nombre(termino):
    conn = conectar_db()
    if not conn: return None
    try:
        cursor = conn.cursor(dictionary=True)
        query_user = "SELECT id_usuario, nombre, email, region FROM usuario WHERE email = %s OR nombre LIKE %s LIMIT 1"
        cursor.execute(query_user, (termino, f"%{termino}%"))
        cliente = cursor.fetchone()
        if cliente:
            query_pedidos = "SELECT COUNT(id_pedido) as total_pedidos FROM pedidos WHERE id_usuario = %s"
            cursor.execute(query_pedidos, (cliente['id_usuario'],))
            cliente['total_pedidos'] = cursor.fetchone()['total_pedidos']
        return cliente
    finally:
        if conn and conn.is_connected(): conn.close()

def obtener_estadisticas_admin():
    conn = conectar_db()
    if not conn: return None
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM contacto_mensajes WHERE leido = 0")
        nuevos_mensajes = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM solicitudes_servicio WHERE estado = 'Pendiente'")
        servicios_pendientes = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM producto WHERE stock < 10 AND activo = 1")
        bajo_stock = cursor.fetchone()[0]
        return { "nuevos_mensajes": nuevos_mensajes, "servicios_pendientes": servicios_pendientes, "bajo_stock": bajo_stock }
    finally:
        if conn and conn.is_connected(): conn.close()

def find_product_id_by_name(product_name, id_vendedor=None):
    conn = conectar_db()
    if not conn: return None
    try:
        cursor = conn.cursor(dictionary=True)
        # 1. Usamos LOWER() en la BD y en Python para ignorar mayúsculas.
        # 2. Comparamos con LIKE para ser más flexibles.
        sql = "SELECT id_producto, nombre, id_vendedor FROM producto WHERE LOWER(nombre) LIKE LOWER(%s)"
        params = [f"%{product_name}%"]

        if id_vendedor:
            sql += " AND id_vendedor = %s"
            params.append(id_vendedor)
        sql += " LIMIT 1"
        cursor.execute(sql, tuple(params))
        producto = cursor.fetchone()
        return producto
    finally:
        if conn and conn.is_connected(): conn.close()

# ======================================================================
# LíGICA DE PREDICCIíN (SARIMAX + PANDAS ACTUALIZADO)
# ======================================================================

def get_prediction_data(product_id):
    try:
        # Usamos db_engine y read_sql normal para evitar warning
        query = """
            SELECT DATE(p.fecha_pedido) as dia, SUM(pp.cantidad) as total_vendido
            FROM pedidos p
            JOIN pedidos_productos pp ON p.id_pedido = pp.id_pedido
            WHERE pp.id_producto = %s
              AND p.estado IN ('Pagado', 'Enviado', 'Entregado')
              AND p.fecha_pedido >= DATE_SUB(NOW(), INTERVAL 6 MONTH)
            GROUP BY DATE(p.fecha_pedido)
            ORDER BY dia ASC;
        """
        sales_data = pd.read_sql(query, db_engine, params=(product_id,))
        
        if sales_data.empty or len(sales_data) < 15:
            return {"success": False, "error": "Datos insuficientes para la predicción (se necesitan al menos 15 dí­as de ventas)."}

        # Procesamiento del DataFrame
        sales_data['dia'] = pd.to_datetime(sales_data['dia'])
        sales_data = sales_data.set_index('dia')
        sales_data['total_vendido'] = pd.to_numeric(sales_data['total_vendido'])
        df_resampled = sales_data.resample('D').sum().fillna(0).astype(float)
        
        order = (1, 1, 1)
        seasonal_order = (1, 1, 1, 7)
        warnings.filterwarnings("ignore") 
        
        model = SARIMAX(df_resampled['total_vendido'],
                        order=order,
                        seasonal_order=seasonal_order,
                        enforce_stationarity=False,
                        enforce_invertibility=False)
        
        model_fit = model.fit(disp=False) 
        forecast = model_fit.forecast(steps=30)
        
        forecast_dates = pd.date_range(start=df_resampled.index.max() + pd.Timedelta(days=1), periods=30).strftime('%Y-%m-%d').tolist()
        forecast_values = [round(val) if val > 0 else 0 for val in forecast.tolist()]

        return {
            "success": True,
            "forecast_labels": forecast_dates,
            "forecast_data": forecast_values,
            "total_forecast": sum(forecast_values) 
        }
    except Exception as e:
        print(f"Error en get_prediction_data: {e}")
        return {"success": False, "error": str(e)}

@app.route("/predict_demand", methods=["POST"])
def predict_demand():
    data = request.json
    product_id = data.get("id_producto")
    if not product_id:
        return jsonify({"error": "Falta id_producto"}), 400

    result = get_prediction_data(product_id)
    
    if result["success"]:
        return jsonify(result)
    else:
        status_code = 404 if "Datos insuficientes" in result["error"] else 500
        return jsonify({"error": result["error"]}), status_code


# ======================================================================
# ENDPOINT PRINCIPAL DEL CHATBOT
# ======================================================================

@app.route("/chat", methods=["POST"])
def chat():
    data = request.json
    mensaje = data.get("message", "").lower()
    permisos = data.get("permisos")
    id_usuario = data.get("userId")
    email_usuario = data.get("email_usuario", "") 
    nombre_usuario = data.get("nombre_usuario", "Invitado")
    
    # URL base
    base_url = "https://bitware.site:5000/"
    
    respuesta, productos_recomendados = "No te entendí.", []
    intencion = clasificar_intencion(mensaje)

    # =================================================================
    # 1. LÓGICA PARA ADMINISTRADORES ('A')
    # =================================================================
    if permisos == 'A':
        if intencion == "saludo":
            alerts = get_proactive_alerts()
            
            if alerts:
                frases_alerta = [
                    f"Hola, Admin **{nombre_usuario}**. Atención: {alerts}",
                    f"Bienvenido de nuevo. Tengo novedades importantes: {alerts}",
                    f"¡Hola! Antes de empezar, revisa esto: {alerts}",
                    f"Saludos, Jefe. Informe de situación: {alerts}",
                    f"Alerta del sistema para **{nombre_usuario}**: {alerts}",
                    f"Atención requerida: {alerts}",
                    f"Reporte de urgencia: {alerts}",
                    f"Hola. He detectado algo que requiere tu mirada: {alerts}",
                    f"Panel de control reporta: {alerts}",
                    f"¡Ojo al dato, Admin! {alerts}"
                ]
                respuesta = random.choice(frases_alerta)
            else:
                frases_ok = [
                    f"Hola, Admin **{nombre_usuario}**. Todo el sistema opera al 100%.",
                    f"¡Bienvenido, **{nombre_usuario}**! No hay alertas pendientes por ahora.",
                    f"Saludos. La plataforma está estable y sin novedades urgentes.",
                    f"¡Qué tal, **{nombre_usuario}**! Todo tranquilo en el servidor hoy.",
                    f"Hola. Todo parece estar en orden. ¿Qué quieres revisar hoy?",
                    f"Sistema nominal. Bienvenido **{nombre_usuario}**.",
                    f"Sin novedades en el frente. ¿Vemos las estadísticas?",
                    f"Todo en verde. ¿Exportamos algún reporte hoy?",
                    f"Hola **{nombre_usuario}**. Base de datos y stock funcionando correctamente.",
                    f"Día tranquilo en la oficina virtual. ¿En qué te ayudo?",
                    f"Bienvenido. El dashboard está limpio de errores."
                ]
                respuesta = random.choice(frases_ok)

        elif intencion == "prediccion_stock":
            match = re.search(r'(?:de|del)\s(.+)', mensaje)
            if not match: match = re.search(r'stock\s(.+)', mensaje)
            termino_busqueda = match.group(1).strip() if match else ""
            
            if not termino_busqueda:
                respuesta = "Claro, dime el nombre del producto que quieres predecir. Ej: 'Predice el stock de RTX 3060'"
            else:
                producto = find_product_id_by_name(termino_busqueda) 
                if not producto:
                    respuesta = f"No encontré el producto '{termino_busqueda}'."
                else:
                    respuesta = f"Analizando la demanda de '{producto['nombre']}', por favor espera..."
                    prediccion_resultado = get_prediction_data(producto['id_producto'])
                    
                    if prediccion_resultado["success"]:
                        total_unidades = prediccion_resultado['total_forecast']
                        respuesta = f"La demanda pronosticada para '{producto['nombre']}' en los próximos 30 días es de **{total_unidades} unidades**."
                    else:
                        respuesta = f"No pude predecir '{producto['nombre']}': {prediccion_resultado['error']}"

        elif intencion == "exportar_ventas":
            respuesta = "Generando reporte GLOBAL de ventas..."
            resultado_url = generar_excel_ventas(id_usuario, base_url, es_admin=True)
            
            if resultado_url == "empty":
                respuesta = "No encontré ventas registradas en el sistema."
            elif resultado_url:
                respuesta = f"✅ Reporte Global Generado.<br><br>👉 <a href='{resultado_url}' target='_blank' style='color: #0d6efd; font-weight: bold;'>Descargar Excel Global</a>"
            else:
                respuesta = "Hubo un error al generar el reporte."
        
        elif intencion == "stats_admin":
            if "total usuarios" in mensaje or "cuantos usuarios" in mensaje:
                conn = conectar_db()
                cursor = conn.cursor()
                cursor.execute("SELECT COUNT(*) as total FROM usuario")
                count = cursor.fetchone()[0]
                conn.close()
                respuesta = f"Actualmente hay <strong>{count}</strong> usuarios registrados en total."
            elif "reporte de ventas hoy" in mensaje or "ventas hoy" in mensaje:
                conn = conectar_db()
                cursor = conn.cursor()
                sql = "SELECT SUM(total) FROM pedidos WHERE estado IN ('Pagado', 'Enviado', 'Entregado') AND DATE(fecha_pedido) = CURDATE()"
                cursor.execute(sql)
                total_hoy = cursor.fetchone()[0] or 0
                conn.close()
                respuesta = f"Los ingresos totales de hoy (pedidos pagados) son: <strong>${total_hoy:,.0f}</strong>."
            else:
                stats = obtener_estadisticas_admin()
                if stats:
                    respuesta = (f"**Resumen Rápido del Sistema:**\n"
                                 f"- **Mensajes Nuevos:** {stats['nuevos_mensajes']}\n"
                                 f"- **Servicios Pendientes:** {stats['servicios_pendientes']}\n"
                                 f"- **Productos con Bajo Stock:** {stats['bajo_stock']}")
                else:
                    respuesta = "No pude obtener las estadísticas en este momento."
        
        elif intencion == "cambiar_estado_pedido":
            match = re.search(r'pedido\s#?(\d+)\s(?:a|como)\s(\w+)', mensaje)
            if match:
                id_pedido, nuevo_estado = match.groups()
                respuesta = cambiar_estado_pedido_db(id_pedido, nuevo_estado.capitalize())
            else:
                respuesta = "Claro. Dime el número de pedido y el nuevo estado. Ej: 'Actualiza el pedido 123 a Enviado'."
        
        elif intencion == "analisis_admin":
            respuesta = get_category_growth_analysis()

        elif intencion == "buscar_cliente_admin":
            match = re.search(r'(?:busca al cliente|datos de cliente|info de|encuentra a|cliente)\s(.+)', mensaje)
            termino_cliente = match.group(1).strip() if match else mensaje
            if termino_cliente:
                cliente = buscar_cliente_por_email_o_nombre(termino_cliente)
                if cliente:
                    respuesta = f"**Cliente Encontrado:**\n- **Nombre:** {cliente['nombre']}\n- **Email:** {cliente['email']}\n- **Región:** {cliente['region'] or 'N/A'}\n- **Pedidos:** {cliente['total_pedidos']}"
                else:
                    respuesta = f"No encontré al cliente '{termino_cliente}'."
            else:
                respuesta = "Dime el nombre o email del cliente que buscas."

        elif intencion == "funciones" or "ayuda" in mensaje:
            respuesta = (
                "**Comandos de Administrador:**\n"
                "* **'Exportar ventas'**: Descarga un reporte global de todas las ventas.\n"
                "* **'Estadísticas'**: Muestra el resumen rápido del sistema.\n"
                "* **'Total usuarios'**: Muestra el conteo total de usuarios registrados.\n"
                "* **'Reporte de ventas hoy'**: Calcula los ingresos del día.\n"
                "* **'Busca al cliente [dato]'**: Encuentra información de un cliente.\n"
                "* **'Actualiza pedido [ID] a [estado]'**: Cambia estado (Ej: 'pedido 105 a Enviado').\n"
                "* **'Predice stock de [producto]'**: Pronóstico de demanda."
            )
        else:
            respuesta = "No entendí ese comando de Admin. Escribe 'ayuda' para ver las opciones."

    # =================================================================
    # 2. LÓGICA PARA VENDEDORES ('V')
    # =================================================================
    elif permisos == 'V':
        if intencion == "saludo":
            frases_vendedor = [
                f"¡Hola, Vendedor **{nombre_usuario}**! ¿Listo para vender más hoy?",
                f"Bienvenido a tu panel, **{nombre_usuario}**. ¿Quieres ver tus métricas o productos?",
                f"¡Saludos! Tu tienda está activa. ¿Necesitas un reporte de ventas?",
                f"Hola **{nombre_usuario}**. ¿En qué puedo ayudarte con tu inventario hoy?",
                f"¡Qué bueno verte, **{nombre_usuario}**! Recuerda revisar tu stock si tienes dudas.",
                f"¡A por todas hoy, **{nombre_usuario}**! ¿Revisamos cómo van tus ingresos?",
                f"Hola. Tus productos están visibles para los clientes. ¿Quieres destacar alguno?",
                f"Bienvenido. ¿Hacemos una predicción de stock para algún producto?",
                f"Saludos emprendedor. ¿Exportamos tu Excel de ventas?",
                f"¡Éxito en las ventas de hoy! Estoy aquí si necesitas datos.",
                f"Hola **{nombre_usuario}**. ¿Consultamos el rendimiento de tu catálogo?"
            ]
            respuesta = random.choice(frases_vendedor)
        
        elif intencion == "prediccion_stock":
            palabras_a_quitar = ["predice stock de", "predice el stock de", "predice stock", "predicción de", "demanda de", "pronostica"]
            termino_busqueda = mensaje
            for frase in palabras_a_quitar:
                if termino_busqueda.startswith(frase):
                    termino_busqueda = termino_busqueda[len(frase):].strip() 
                    break
            
            if not termino_busqueda:
                respuesta = "Claro, dime el nombre de tu producto que quieres predecir. Ej: 'Predice el stock de AdoLuche'"
            else: 
                producto = find_product_id_by_name(termino_busqueda, id_vendedor=id_usuario) 
                if not producto:
                    respuesta = f"No encontré el producto '{termino_busqueda}' en tu inventario."
                else:
                    respuesta = f"Analizando la demanda de '{producto['nombre']}', por favor espera..."
                    prediccion_resultado = get_prediction_data(producto['id_producto'])
                    
                    if prediccion_resultado["success"]:
                        total_unidades = prediccion_resultado['total_forecast']
                        respuesta = f"La demanda pronosticada para tu producto '{producto['nombre']}' en los próximos 30 días es de **{total_unidades} unidades**."
                    else:
                        respuesta = f"No pude predecir '{producto['nombre']}': {prediccion_resultado['error']}"
        
        elif intencion == "exportar_ventas":
            respuesta = "Generando tu reporte de ventas seguro, dame un momento..."
            resultado_url = generar_excel_ventas(id_usuario, base_url)
            
            if resultado_url == "empty":
                respuesta = "Revisé tus registros y no encontré ventas pagadas o finalizadas para exportar."
            elif resultado_url:
                respuesta = f"¡Listo! He generado tu archivo Excel.<br><br><a href='{resultado_url}' target='_blank' style='color: #0d6efd; font-weight: bold;'>Descargar Reporte</a>"
            else:
                respuesta = "Hubo un error interno al generar el archivo. Por favor intenta más tarde."

        elif intencion == "stats_admin" or intencion == "stock_admin" or "productos" in mensaje or "ventas" in mensaje:
             conn = conectar_db()
             sql_prod = "SELECT COUNT(*) as num_productos, SUM(stock) as total_stock FROM producto WHERE id_vendedor = %s"
             cursor = conn.cursor(dictionary=True)
             cursor.execute(sql_prod, (id_usuario,))
             result_prod = cursor.fetchone()
             num_productos = result_prod.get('num_productos') or 0
             total_stock = result_prod.get('total_stock') or 0
             
             sql_ventas = "SELECT COUNT(DISTINCT p.id_pedido) as num_ventas, SUM(pp.cantidad * pp.precio_unitario) as total_revenue FROM producto pr JOIN pedidos_productos pp ON pr.id_producto = pp.id_producto JOIN pedidos p ON pp.id_pedido = p.id_pedido WHERE pr.id_vendedor = %s AND p.estado IN ('Pagado', 'Enviado', 'Entregado')"
             cursor.execute(sql_ventas, (id_usuario,))
             result_ventas = cursor.fetchone()
             num_ventas = result_ventas.get('num_ventas') or 0
             total_revenue = result_ventas.get('total_revenue') or 0
             conn.close()

             if "ventas" in mensaje:
                 respuesta = f"Hasta ahora, has realizado <strong>{num_ventas}</strong> ventas, generando un total de <strong>${total_revenue:,.0f}</strong>."
             else:
                 respuesta = f"Actualmente tienes <strong>{num_productos}</strong> productos listados, con un stock total de <strong>{total_stock}</strong> unidades."
        
        elif intencion == "funciones" or "ayuda" in mensaje:
            respuesta = (
                "**Comandos de Vendedor:**\n"
                "* **'Exportar ventas'**: Descarga un Excel seguro con tus transacciones.\n"
                "* **'Mis ventas'**: Muestra el total de ventas e ingresos.\n"
                "* **'Mis productos'**: Muestra tu inventario y stock.\n"
                "* **'Predice stock de [mi producto]'**: Pronóstico de demanda."
            )
        else:
            respuesta = "No entendí ese comando. Escribe 'ayuda' para ver tus opciones."

    # =================================================================
    # 3. LÓGICA PARA CLIENTES / INVITADOS ('U' o None)
    # =================================================================
    else:
        if intencion == "prediccion_stock" or intencion == "exportar_ventas":
            respuesta = "Lo siento, esa función es exclusiva para Vendedores y Administradores."
            
        elif not id_usuario and intencion in ["pedido", "solicitar_devolucion", "solicitar_notificacion", "actualizar_direccion"]:
            respuesta = "Para esa función, primero debes **iniciar sesión** en tu cuenta."
        
        elif intencion == "saludo":
            if id_usuario:
                frases_cliente = [
                    f"¡Hola de nuevo, **{nombre_usuario}**! ¿Buscas algo especial hoy?",
                    f"Bienvenido a Bitware, **{nombre_usuario}**. ¿En qué te ayudo?",
                    f"¡Qué tal, **{nombre_usuario}**! ¿Quieres revisar el estado de algún pedido?",
                    f"Hola **{nombre_usuario}**. Estoy aquí para ayudarte a encontrar el mejor hardware.",
                    f"Saludos. Recuerda que puedes pedirme comparar productos o ver tu historial.",
                    f"¡Qué gusto verte! ¿Te ayudo a rastrear tu última compra?",
                    f"Hola **{nombre_usuario}**. ¿Necesitas ayuda con alguna devolución o duda?",
                    f"Bienvenido. ¿Buscas actualizar tu setup hoy?",
                    f"Hola. Si necesitas cambiar tu dirección, solo dímelo.",
                    f"Saludos **{nombre_usuario}**. ¿Vemos qué hay de nuevo en stock?",
                    f"Aquí estoy. Pregúntame por 'mi pedido' o busca un componente."
                ]
                respuesta = random.choice(frases_cliente)
            else:
                frases_invitado = [
                    "¡Hola! Bienvenido a Bitware. Inicia sesión para sacar el máximo provecho.",
                    "¡Bienvenido! Soy tu asistente virtual. Escribe 'ayuda' para ver qué puedo hacer.",
                    "Hola. ¿Buscas algún componente en específico? Puedo ayudarte a buscar.",
                    "¡Saludos! Explora nuestro catálogo o pregúntame por horarios y métodos de pago.",
                    "Bienvenido al futuro del hardware. ¿Te ayudo a encontrar algo?",
                    "Hola viajero. ¿Necesitas comparar precios de componentes?",
                    "¡Buenas! Soy el bot de Bitware. Pregúntame lo que necesites.",
                    "Bienvenido. ¿Te interesa saber nuestros métodos de pago?",
                    "Hola. Puedo ayudarte a cotizar o comparar productos. ¡Prueba!",
                    "Saludos. Inicia sesión para ver tus pedidos, o pregúntame libremente.",
                    "¡Hola! ¿Buscas soporte o ventas? Estoy para servirte."
                ]
                respuesta = random.choice(frases_invitado)
            
        elif intencion == "funciones" or "ayuda" in mensaje:
            respuesta_basica = (
                "* **'Busca [producto]'**: Para encontrar productos (Ej: 'Busca RTX 3060').\n"
                "* **'Compara [A] con [B]'**: Muestra una comparativa de precios (Ej: 'Compara Intel i5 con Ryzen 5').\n"
                "* **'Horarios de atención'**: Muestra los horarios de la tienda.\n"
                "* **'Métodos de pago'**: Lista las formas de pago aceptadas.\n"
                "* **'Soporte'**: Te indica cómo contactar a soporte técnico."
            )
            
            if not id_usuario:
                respuesta = (
                    "**Hola, Invitado. Esto es lo que puedes hacer:**\n"
                    f"{respuesta_basica}\n"
                    "\n¡**Inicia sesión** para ver tu pedido, actualizar tu dirección y más!"
                )
            else:
                respuesta = (
                    f"**Hola, {nombre_usuario}. ¡Puedes pedirme todo esto!:**\n\n"
                    "**SOBRE TUS PEDIDOS:**\n"
                    "* **'Estado de mi pedido'**: Revisa dónde está tu última compra.\n"
                    "* **'Quiero devolver [motivo]'**: Inicia una solicitud de devolución para tu último pedido.\n\n"
                    "**SOBRE PRODUCTOS:**\n"
                    "* **'Avísame de [producto]'**: Te notificaré por email cuando un producto sin stock vuelva a estar disponible.\n"
                    f"{respuesta_basica}\n\n"
                    "**SOBRE TU CUENTA:**\n"
                    "* **'Actualizar mi dirección a [dirección]'**: Cambia tu dirección de envío principal."
                )
        
        elif intencion == "producto":
            productos_recomendados = recomendar_productos()
            respuesta = "¡Claro! Aquí tienes algunas recomendaciones:"

        elif intencion == "busqueda_producto":
            frases_activadoras = ["buscar producto", "busca producto", "encontrar producto"]
            if mensaje in frases_activadoras:
                respuesta = "¿Qué producto te gustaría buscar?"
            else:
                termino_busqueda = mensaje
                palabras_clave_iniciales = ["busca", "búscame", "encuentra", "tienes", "precio de", "cotiza", "buscar", "quiero buscar"]
                for palabra in palabras_clave_iniciales:
                    if termino_busqueda.startswith(palabra + " "):
                        termino_busqueda = termino_busqueda[len(palabra)+1:].strip()
                        break 
                if not termino_busqueda:
                    respuesta = "¿Qué producto te gustaría buscar?"
                else:
                    productos_encontrados = buscar_productos_por_nombre(termino_busqueda)
                    if productos_encontrados:
                        respuesta = f"Encontré esto relacionado con **'{termino_busqueda}'**:"
                        productos_recomendados = productos_encontrados
                    else:
                        respuesta = f"Lo siento, no encontré nada relacionado con **'{termino_busqueda}'**."
        
        elif intencion == "pedido":
            pedido = estado_ultimo_pedido(id_usuario)
            respuesta = f"Tu último pedido es el #{pedido['id_pedido']} y su estado es: **{pedido['estado']}**." if pedido else "Aún no tienes pedidos."
        
        elif intencion == "solicitar_devolucion":
            check = solicitar_devolucion_db(id_usuario)
            respuesta = check["mensaje"]
        
        elif intencion == "solicitar_notificacion":
            match = re.search(r'(?:avísame de|notifícame de|disponible)\s(?:el|la)\s(.+)', mensaje)
            producto_nombre = match.group(1).strip() if match else ""
            respuesta = solicitar_notificacion_db(id_usuario, email_usuario, producto_nombre)
        
        elif intencion == "comparar_productos":
            match = re.search(r'compara\s(.+)\s(?:con|y)\s(.+)', mensaje)
            if match:
                p1, p2 = match.groups()
                respuesta = comparar_productos_db(p1.strip(), p2.strip())
            else:
                respuesta = "Dime los dos productos que quieres comparar. Ej: 'Compara RTX 3060 con RX 6600'."
        
        elif intencion == "actualizar_direccion":
            match = re.search(r'dirección\s(?:a|es)\s(.+)', mensaje)
            nueva_dir = match.group(1).strip() if match else ""
            respuesta = actualizar_direccion_db(id_usuario, nueva_dir) if nueva_dir else "Dime cuál es tu nueva dirección. Ej: 'Actualizar mi dirección a Calle Falsa 123'."

        elif intencion == "horarios":
            respuesta = "Nuestros horarios de atención son de **Lunes a Viernes de 9:00 a 18:00 hrs**."
        
        elif intencion == "pagos":
            respuesta = "Aceptamos pagos a través de **Webpay (Tarjetas de Crédito/Débito)**"
        
        elif intencion == "soporte":
            respuesta = "Para soporte técnico, visita nuestra sección de **Ayuda** o envíanos un mensaje desde **Soporte** en el pie de página."
        
        else:
            respuesta = "Lo siento, no entendí. Puedes pedirme que **busque un producto** o que revise el **estado de tu pedido**."

    # ESTE ES EL ÚNICO RETURN QUE DEBE HABER AL FINAL
    return jsonify({"respuesta": respuesta, "productos": productos_recomendados})

if __name__ == "__main__":
    # --- CONFIGURACIÓN SSL (HTTPS) AUTOMÁTICA ---
    cert_path = '/etc/letsencrypt/live/bitware.site/fullchain.pem'
    key_path = '/etc/letsencrypt/live/bitware.site/privkey.pem'

    if os.path.exists(cert_path) and os.path.exists(key_path):
        print("✅ MODO SEGURO: Certificados SSL encontrados. Iniciando HTTPS...")
        # Contexto SSL para que Flask sirva en https://bitware.site:5000
        context = (cert_path, key_path)
        app.run(host='0.0.0.0', port=5000, ssl_context=context, debug=True)
    else:
        print("⚠️ MODO INSEGURO: Certificados no encontrados. Iniciando HTTP simple.")
        app.run(host='0.0.0.0', port=5000, debug=True)
